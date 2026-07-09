from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Union

from openpyxl import load_workbook


PRICE_COLUMNS = {
    "material_price": 7,
    "auxiliary_price": 8,
    "labor_price": 9,
}


Price = Union[int, float]


def read_prices(workbook_path: Path) -> dict[str, dict[str, Price]]:
    workbook = load_workbook(workbook_path, data_only=False)
    worksheet = workbook.active
    prices: dict[str, dict[str, Price]] = {}
    for row_index in range(4, worksheet.max_row + 1):
        item_name = worksheet.cell(row_index, 3).value
        if not isinstance(item_name, str) or not item_name.strip():
            continue
        prices[item_name.strip()] = {
            key: normalize_price(worksheet.cell(row_index, column).value)
            for key, column in PRICE_COLUMNS.items()
        }
    return prices


def normalize_price(value: Any) -> Price:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return normalize_number(float(value))
    if isinstance(value, str):
        trimmed = value.strip()
        if not trimmed or trimmed.startswith("="):
            return 0
        return normalize_number(float(trimmed))
    raise ValueError(f"无法读取单价：{value!r}")


def normalize_number(value: float) -> Price:
    rounded = round(value, 2)
    if rounded.is_integer():
        return int(rounded)
    return rounded


def sync_rules_json(rules_path: Path, prices: dict[str, dict[str, Price]]) -> tuple[list[dict[str, Any]], int]:
    rules: list[dict[str, Any]] = json.loads(rules_path.read_text(encoding="utf-8"))
    updated = 0
    for rule in rules:
        item_name = rule.get("item_name")
        if item_name not in prices:
            continue
        price = prices[item_name]
        next_unit_price = normalize_number(price["material_price"] + price["auxiliary_price"] + price["labor_price"])
        before = (
            rule.get("material_price", 0),
            rule.get("auxiliary_price", 0),
            rule.get("labor_price", 0),
            rule.get("unit_price", 0),
        )
        rule["material_price"] = price["material_price"]
        rule["auxiliary_price"] = price["auxiliary_price"]
        rule["labor_price"] = price["labor_price"]
        rule["unit_price"] = next_unit_price
        after = (
            rule["material_price"],
            rule["auxiliary_price"],
            rule["labor_price"],
            rule["unit_price"],
        )
        if after != before:
            updated += 1
    rules_path.write_text(f"{json.dumps(rules, ensure_ascii=False, indent=2)}\n", encoding="utf-8", newline="\n")
    return rules, updated


def sync_quote_mapping_defaults(source_path: Path, rules: list[dict[str, Any]]) -> int:
    source = source_path.read_text(encoding="utf-8")
    updated = 0
    for rule in rules:
        item_name = re.escape(str(rule["item_name"]))
        metric = re.escape(str(rule["metric"]))
        unit = re.escape(str(rule["unit"]))
        material = format_number(rule["material_price"])
        auxiliary = format_number(rule["auxiliary_price"])
        labor = format_number(rule["labor_price"])
        unit_price = format_number(rule["unit_price"])
        pattern = re.compile(rf'(quoteRule\("{item_name}",\s*"{metric}",\s*"{unit}",\s*)[-0-9.]+,\s*[-0-9.]+,\s*[-0-9.]+(?P<tail>[^\n]*?\))')
        source, replacements = pattern.subn(lambda match: f"{match.group(1)}{material}, {auxiliary}, {labor}{sync_unit_price_tail(match.group('tail'), unit_price)}", source, count=1)
        if replacements:
            updated += replacements
    source_path.write_text(source, encoding="utf-8", newline="\n")
    return updated


def sync_unit_price_tail(tail: str, unit_price: str) -> str:
    return re.sub(r"(,\s*undefined,\s*)[-0-9.]+", rf"\g<1>{unit_price}", tail)


def format_number(value: Price) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync checked quote prices into default quote rules.")
    parser.add_argument("--xlsx", default=r"D:\Desktop\报价规则单价核对表.xlsx")
    parser.add_argument("--rules", default="quote-rules-apartment-current.json")
    parser.add_argument("--source", default=r"apps\web\lib\quote-mapping.ts")
    args = parser.parse_args()

    prices = read_prices(Path(args.xlsx))
    rules, json_updates = sync_rules_json(Path(args.rules), prices)
    source_updates = sync_quote_mapping_defaults(Path(args.source), rules)
    print(f"价格表读取 {len(prices)} 项；JSON 更新 {json_updates} 项；默认源码同步 {source_updates} 项。")


if __name__ == "__main__":
    main()
