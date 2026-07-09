from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


GROUPS = [
    (
        "墙顶地/湿区",
        {
            "墙面界面剂处理",
            "墙面批嵌",
            "墙面乳胶漆",
            "轻钢龙骨平顶",
            "双眼皮/边吊吊顶",
            "石膏线吊顶",
            "顶面批嵌",
            "顶面乳胶漆",
            "厨房卫生间集成吊顶",
            "地面找平",
            "地面砖铺贴(750X1500)",
            "地面瓷砖",
            "墙面瓷砖",
            "瓷砖加工费",
            "美缝",
            "墙面贴瓷砖(600X1200)",
            "墙地面防漏处理",
            "窗台石铺贴",
            "楼梯踏步铺贴",
        },
    ),
    (
        "全屋拆改/其他工程",
        {
            "砌砖墙",
            "砌120厚砖墙",
            "砌240厚砖墙",
            "现浇钢筋混凝土楼板",
            "拆改及拆墙",
            "外墙批嵌以及修补",
            "砖墙门窗洞过梁",
            "水泥墙开槽",
            "打混凝土过梁孔",
            "厨房、卫生间排污管包隔音棉",
            "补线、管槽及零星修补",
        },
    ),
    (
        "水电/项目服务",
        {
            "强电插座",
            "开关",
            "灯位",
            "筒灯/射灯",
            "设备专线",
            "弱电点位",
            "强电线管",
            "弱电线管",
            "强电箱",
            "弱电箱",
            "分配电箱",
            "给水点",
            "热水点",
            "排水点",
            "给水管",
            "排水管",
            "强电布线",
            "弱电布线",
            "水路布管",
            "材料搬运费",
            "垃圾清运费",
            "墙地面砖现场保护",
            "全屋保洁",
        },
    ),
    (
        "门窗/定制",
        {
            "背景墙",
            "入户门",
            "室内门",
            "卫生间门",
            "厨房推拉门",
            "厨房推拉门双包套",
            "阳台推拉门",
            "阳台推拉门双包套",
            "铝合金封门窗",
            "橱柜",
            "全屋定制",
        },
    ),
    (
        "洁具/灯饰",
        {
            "马桶",
            "蹲坑",
            "浴室柜",
            "淋浴隔断",
            "玻璃淋浴房",
            "淋浴隔断安装",
            "花洒",
            "卫浴五件套",
            "全屋插座开关",
            "全屋灯饰",
        },
    ),
    ("窗帘/收口", {"窗帘", "窗台石", "暗窗帘箱", "楼梯扶手", "栏杆/护栏"}),
]


def category(item_name: str) -> str:
    for title, item_names in GROUPS:
        if item_name in item_names:
            return title
    return "其他规则"


def export_quote_rule_check_xlsx(rules_path: Path, output_path: Path) -> None:
    rules: list[dict[str, Any]] = json.loads(rules_path.read_text(encoding="utf-8"))
    wb = Workbook()
    ws = wb.active
    ws.title = "报价规则单价核对"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.35, header=0.2, footer=0.2)

    headers = ["序号", "分类", "清单项", "取数指标", "单位", "适用空间", "主材", "辅材", "人工", "汇总单价", "备注"]
    ws.merge_cells("A1:K1")
    ws["A1"] = "报价规则单价核对表"
    ws["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws["A2"] = "来源"
    ws["B2"] = str(rules_path)
    ws["C2"] = "规则数"
    ws["D2"] = len(rules)
    ws["E2"] = "口径"
    ws["F2"] = "系统默认规则"

    for column, header in enumerate(headers, 1):
        cell = ws.cell(3, column, header)
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D9E2F3")
    for index, rule in enumerate(rules, 1):
        row_index = index + 3
        values = [
            index,
            category(rule["item_name"]),
            rule["item_name"],
            rule["metric"],
            rule["unit"],
            "、".join(rule.get("space_types") or ["全部"]),
            rule.get("material_price", 0),
            rule.get("auxiliary_price", 0),
            rule.get("labor_price", 0),
            f"=G{row_index}+H{row_index}+I{row_index}",
            "",
        ]
        for column, value in enumerate(values, 1):
            cell = ws.cell(row_index, column, value)
            cell.font = Font(name="Microsoft YaHei", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
        for column in range(7, 11):
            ws.cell(row_index, column).number_format = "0.00"

    widths = [6, 16, 24, 28, 8, 46, 10, 10, 10, 12, 20]
    for column, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(column)].width = width
    for row_index in range(2, ws.max_row + 1):
        ws.row_dimensions[row_index].height = 24 if row_index == 3 else 34
    ws.auto_filter.ref = f"A3:K{ws.max_row}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--rules", default="quote-rules-apartment-current.json")
    parser.add_argument("--output", default=r"D:\Desktop\报价规则单价核对表.xlsx")
    args = parser.parse_args()
    export_quote_rule_check_xlsx(Path(args.rules), Path(args.output))


if __name__ == "__main__":
    main()
